#!/usr/bin/env python
# coding: utf-8

import pdfplumber
import re
import openpyxl
import os
import tkinter as tk
from tkinter import font as tkFont
from tkinter import filedialog

from datetime import datetime


def GSTR3B():
    dt = datetime.now().strftime("%Y_%m_%d-%I_%M_%S_%p")

    directory = filedialog.askdirectory()

    workbook = openpyxl.Workbook() 

    worksheet = workbook.active
    worksheet2 = workbook.create_sheet('payment')

    titles = ['Taxable Supplies', 'Zero Rated', 'Nil/ Exempted', 'RCM', 'Non-GST', 'Supplies to URP', 'Supplies to CTP', 'Supplies to UIN', 'ITC Available',
             'ITC Reversed', 'Net ITC', 'Ineligible ITC', 'Interest', 'Late Fees', 'TDS', 'TCS']

    titles2 = ['IGST - Other than RCM',  'CGST - Other than RCM', 'SGST - Other than RCM', 'Cess - Other than RCM',
               'IGST - RCM','CGST - RCM', 'SGST - RCM', 'CESS - RCM', ]

    r = 3
    for title in titles:
        worksheet.cell(row = r, column = 1).value = title
        r = r + 1

    r = 3
    for title2 in titles2:
        worksheet2.cell(row = r, column = 1).value = title2
        r = r + 1   

    pay_re = re.compile(r'(-|\d*\.?\d*)\s(-|\d*\.?\d*)\s(-|\d*\.?\d*)\s(-|\d*\.?\d*)\s(-|\d*\.?\d*)\s(-|\d*\.?\d*)\s(-|\d*\.?\d*)$')

    c = 2
    cc = 2
    col_counter = 2
    col_counter2 = 2

    for filename in os.listdir(directory):
        rc = 2
        if filename.endswith(".pdf"):
            titles = ['Taxable Value', 'IGST', 'CGST', 'SGST', 'CESS']
            for title in titles:
                worksheet.cell(row = 2, column = c).value = title
                worksheet.cell(row = 1, column = c).value = filename[-10:].replace('.pdf', '')
                c = c + 1

            titles2 = ['Tax payable', 'IGST ', 'CGST', 'SGST', 'CESS', 'Cash', 'Interest', 'Late Fees']
            for title2 in titles2:
                worksheet2.cell(row = 2, column = cc).value = title2
                worksheet2.cell(row = 1, column = cc).value = filename[-10:].replace('.pdf', '')
                cc = cc + 1

            with pdfplumber.open(os.path.join(directory, filename)) as pdf:
                pages = pdf.pages
                for page in pages:
                    text = page.extract_text()

                    for line in text.split('\n'):
                        if line.startswith('(a)'):
                            col_count = col_counter
                            i = -1
                            for i in range(-5, 0):
                                store = line.split()[i]
                                if len(store) ==1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 3, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('(b)'):
                            i = -1
                            col_count = col_counter
                            for i in range(-5, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 4, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('(c)'):
                            i = -1
                            col_count = col_counter
                            for i in range(-5, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 5, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('(d)'):
                            i = -1
                            col_count = col_counter
                            for i in range(-5, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 6, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('(e)'):
                            i = -1
                            col_count = col_counter
                            for i in range(-5, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 7, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('Supplies made to Un'):
                            col_count = col_counter
                            i = -1
                            for i in range(-2, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 8, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('Supplies made to Co'):
                            col_count = col_counter
                            i = -1
                            for i in range(-2, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 9, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('Supplies made to UI'):
                            col_count = col_counter
                            i = -1
                            for i in range(-2, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 10, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('(A) I'):
                            col_count = col_counter + 1
                            i = -1
                            for i in range(-4, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 11, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('(B) I'):
                            col_count = col_counter + 1
                            i = -1
                            for i in range(-4, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row=12, column=col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('(C)'):
                            col_count = col_counter + 1
                            i = -1
                            for i in range(-4, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row=13, column=col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('(D)'):
                            col_count = col_counter + 1
                            i = -1
                            for i in range(-4, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row=14, column=col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('Interest'):
                            col_count = col_counter + 1
                            i = -1
                            for i in range(-4, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 15, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif line.startswith('Late'):
                            col_count = col_counter + 1
                            i = -1
                            for i in range(-4, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row = 16, column = col_count).value = float(store)
                                col_count = col_count + 1

                        elif re.compile(r'^TDS').search(line):
                            col_count = col_counter + 1
                            i = -1
                            for i in range(-3, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row=17, column=col_count).value = float(store)
                                col_count = col_count + 1

                        elif re.compile(r'^TCS').search(line):
                            col_count = col_counter + 1
                            i = -1
                            for i in range(-3, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet.cell(row=18, column=col_count).value = float(store)
                                col_count = col_count + 1

                        elif pay_re.search(line):
                            rc = rc + 1
                            col_count = col_counter2
                            i = -1
                            for i in range(-8, 0):
                                store = line.split()[i]
                                if len(store) == 1 and store == "-":
                                    store = store.replace("-", "0")
                                worksheet2.cell(row = rc, column = col_count).value = float(store)
                                col_count = col_count + 1

        col_counter = col_counter + 5
        col_counter2 = col_counter2 + 8

    workbook.save("GSTR3B_converted_" + dt + ".xlsx")

def GSTR1():
    dt = datetime.now().strftime("%Y_%m_%d-%I_%M_%S_%p")

    directory = filedialog.askdirectory()

    workbook = openpyxl.Workbook() 

    worksheet = workbook.active

    in_line_re = re.compile(r'(^-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)$')
    in_line_re2 = re.compile(r'(^-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)$')
    in_line_re3 = re.compile(r'(^-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)\s(-?\d*\.?\d+|-)$')

    # in_line_re = re.compile(r'(^\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)$')
    # in_line_re2 = re.compile(r'(^\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)$')
    # in_line_re3 = re.compile(r'(^\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)\s(\d*\.?\d*)$')

    row_heads = ['B2B', 'CDNR', 'B2CS','Advances', 'Adjustment of Advances', 'HSN', 'Amended B2B', 'Amended CDNR', 'Amended B2CS',
                'Amended Advances', 'Amended Adjustment of Advances','B2CL', 'CDNUR', 'Amended B2CL', 'Amended CDNUR', 'Export',
                 'Nil/Exem/NonGST (Ignore Column Heads)', 'Documents Issued/Cancelled/Net (Ignore Column Heads)',
                'Amended Exports']
    a = 3
    for row_head in row_heads:
        worksheet.cell(row = a, column = 1).value = row_head
        a = a + 1

    c = 2
    col_counter = 2

    for filename in os.listdir(directory):
        if filename.endswith(".pdf"):
            r = 3
            r1 = 14
            r2 = 18
            with pdfplumber.open(os.path.join(directory, filename)) as pdf:
                col_heads = ['Invoice Value','Taxable Value', 'IGST', 'CGST', 'SGST', 'CESS']
                for col_head in col_heads:
                    worksheet.cell(row = 1, column = c).value = filename[-10:].replace('.pdf', '')
                    worksheet.cell(row = 2, column = c).value = col_head
                    c = c + 1
                pages = pdf.pages
                for page in pages:
                    text = page.extract_text()

                    for line in text.split('\n'):
                        col_count = col_counter

                        if in_line_re.search(line):
                            for i in range(1, 7):
                                value_store = line.split()[i]
                                if len(value_store)==1 and value_store=="-":
                                    value_store = value_store.replace("-", "0")
                                worksheet.cell(row=r, column=col_count).value = float(value_store)
                                col_count = col_count + 1
                            r = r + 1

                        if in_line_re2.search(line):
                            for i in range(1, 5):
                                value_store = line.split()[i]
                                if len(value_store)==1 and value_store=="-":
                                    value_store = value_store.replace("-", "0")
                                worksheet.cell(row=r1, column=col_count).value = float(value_store)
                                if i == 3:
                                    col_count = col_count + 3
                                else:
                                    col_count = col_count + 1
                            r1 = r1 + 1

                        if in_line_re3.search(line):
                            for i in range(1, 4):
                                value_store = line.split()[i]
                                if len(value_store)==1 and value_store=="-":
                                    value_store = value_store.replace("-", "0")
                                worksheet.cell(row=r2, column=col_count).value = float(value_store)
                                col_count = col_count + 1
                            r2 = r2 + 1

                col_counter = col_counter + 6

    workbook.save('GSTR1_converted_'+ dt +'.xlsx')

def resource_path(relative_path):
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

bg = resource_path("bg.png")

WIDTH = 800
HEIGHT = 600

root = tk.Tk()
root.title("GST returns PDF to Excel by Naresh")

hel = tkFont.Font(family='Helvetica', size=16, weight='bold')

canvas = tk.Canvas(root, height=HEIGHT, width=WIDTH)
canvas.pack()

background_image = tk.PhotoImage(file=bg)
background_label = tk.Label(root, image=background_image)
background_label.place(relwidth=1, relheight=1)

browsebutton = tk.Button(root, text="Convert GSTR 3B", command=GSTR3B, bg='#6699ff', font=hel, fg='white')
browsebutton.place(x=50, y=50, height=50, width=200)

browsebutton = tk.Button(root, text="Convert GSTR 1", command=GSTR1, bg='#6699ff', font=hel, fg='white')
browsebutton.place(x=50, y=130, height=50, width=200)

label = tk.Label(background_label, font=('tahoma', 11),text="Developed by Naresh", fg='blue')
label.place(x=50, y=500, height=50, width=200)

label = tk.Label(background_label, bg="#99ccff", font=('arial', 10),text="You can convert multiple PDF's to one sheet with this tool\nAll you need to do is click the button and\n "
                                                            "select the folder where PDF's are stored\n"
                                                            "That's it, in few seconds you'll have your converted excel \n"
                                                                         "in the folder where this tool is stored\n"
                                                                         "Note - Let the name of PDF files be as it was when downloaded"
                                                                         "\nfrom GST portal (means don't change file names)", fg='#003300')
label.place(x=20, y=250, height=120, width=400)

root.mainloop()